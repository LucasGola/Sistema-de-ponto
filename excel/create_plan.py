import openpyxl as opxl


class Create():
    def __init__(self):
        # instancia a classe que tem uma planilha por padrão chamada 'Sheet1'
        self.arquivo = opxl.Workbook()
        self.arquivo.active

    def active_plan(self, plan):
        # para "ligar" a planilha 'Sheet1' que também é a planilha atuamente ativa
        plan.active

    def set_name(self, plan, nome):
        print(plan)
        plan.title = nome  # Renomeia a planilha

    def create_plan(self, nome):
        # Cria uma nova planilha no arquivo
        self.arquivo.create_sheet(nome)

    def create_set_position(self, nome, position):
        self.arquivo.create_sheet(nome, position)

    def show_plans(self, arquivo):
        print(arquivo.sheetnames)


arquivo = Create()
plan1 = arquivo.active_plan
arquivo.set_name(plan1, "First Plan")
arquivo.create_plan("Second Plan")
arquivo.create_set_position("Tird in First")
arquivo.show_plans(arquivo)
