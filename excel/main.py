from openpyxl import Workbook, load_workbook


def inicializate():
    # instancia a classe que tem uma planilha por padrão chamada 'Sheet1'
    arquivo = Workbook()
    return arquivo


def active(arquivo):
    plan = arquivo.active
    return plan


def set_name(plan, name):
    plan.title = name
    return plan


def create_plan(arquivo, nome):
    # Cria uma nova planilha no arquivo
    plan = arquivo.create_sheet(nome)
    return plan


def create_set_position(arquivo, nome, position):
    plan = arquivo.create_sheet(nome, position)
    return plan


def create_set_position(arquivo, nome, position):
    arquivo.create_sheet(nome, position)


def show_plans(arquivo):
    print(arquivo.sheetnames)


def insert_cell(plan, cell, value):
    plan[cell] = value


def insert_lines(lines, clumns, plan):
    lista = []
    final = []
    l = 0
    c = 0
    for i in range(0, lines):
        for j in range(0, clumns):
            lista.append(
                str(input(f'Insira um valor na linha {l}, coluna {c}.')))
            c += 1
        final.append(tuple(lista))
        c = 0
        l += 1
    l = 0

    for linha in final:
        plan.append(linha)


def set_calc(plan, cell, value):
    plan[cell] = value


def read_cell(plan, cell):
    cl = plan[cell]
    print(cl.value)


def read_all_lines(plan):
    max_lines = plan.max_row
    max_columns = plan.max_column
    for i in range(1, max_lines + 1):
        for j in range(1, max_columns + 1):
            print(plan.cell(row=i, column=j).value, end='-')


def save(arquive, name):
    arquive.save(f'{name}.xlsx')


def open_plan(arquive, path):
    arquive = load_workbook(path)
    return arquive


""" def copy_plan(arquive, plan1, name_to_save):
    origin = arquive.get_sheet_by_name(plan1)
    arquive.copy_worksheet(origin)
    arquive.save(f'{name_to_save}.xlsx') """
